#!/usr/bin/env python3
"""
Recalculate an .xlsx workbook with LibreOffice headless and scan for Excel errors.

Usage:
  python recalc.py <model.xlsx> [timeout_seconds]

Output JSON contract:
  {
    "status": "success" | "errors_found" | "libreoffice_missing" | "recalc_failed",
    "workbook": "...",
    "total_errors": 0,
    "total_formulas": 0,
    "error_summary": {"#REF!": {"count": 1, "locations": ["DCF!B12"]}},
    "recalc_command": [...]
  }
"""

from __future__ import annotations

import json
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

ERROR_TOKENS = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#NULL!", "#NUM!", "#N/A")


def find_soffice() -> str | None:
    for exe in ("soffice", "libreoffice"):
        path = shutil.which(exe)
        if path:
            return path
    # macOS common location
    mac = Path("/Applications/LibreOffice.app/Contents/MacOS/soffice")
    if mac.exists():
        return str(mac)
    return None


def count_formulas(path: Path) -> int:
    wb = load_workbook(path, data_only=False, read_only=True)
    total = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    total += 1
    wb.close()
    return total


def scan_errors(path: Path) -> dict:
    wb = load_workbook(path, data_only=False, read_only=True)
    summary: dict[str, dict[str, object]] = {}
    total = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value in ERROR_TOKENS:
                    item = summary.setdefault(value, {"count": 0, "locations": []})
                    item["count"] = int(item["count"]) + 1
                    item["locations"].append(f"{ws.title}!{cell.coordinate}")
                    total += 1
    wb.close()
    return {"total_errors": total, "error_summary": summary}


def recalc_with_libreoffice(path: Path, timeout: int) -> tuple[bool, list[str], str]:
    soffice = find_soffice()
    if not soffice:
        return False, [], "LibreOffice/soffice not found on PATH"

    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)
        src = tmp / path.name
        out_dir = tmp / "out"
        out_dir.mkdir()
        shutil.copy2(path, src)
        cmd = [
            soffice,
            "--headless",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(out_dir),
            str(src),
        ]
        proc = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
        if proc.returncode != 0:
            return False, cmd, (proc.stderr or proc.stdout or "LibreOffice conversion failed")
        converted = out_dir / path.name
        if not converted.exists():
            # LibreOffice may keep the original stem but normalize extension.
            matches = list(out_dir.glob("*.xlsx"))
            if not matches:
                return False, cmd, f"No converted xlsx produced. stdout={proc.stdout} stderr={proc.stderr}"
            converted = matches[0]
        shutil.copy2(converted, path)
        return True, cmd, proc.stdout.strip()


def main() -> int:
    if len(sys.argv) < 2:
        print(json.dumps({"status": "usage_error", "message": "Usage: python recalc.py <model.xlsx> [timeout_seconds]"}, indent=2))
        return 2
    path = Path(sys.argv[1]).resolve()
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30
    if not path.exists():
        print(json.dumps({"status": "missing_file", "workbook": str(path)}, indent=2))
        return 2

    total_formulas = count_formulas(path)
    ok, cmd, message = recalc_with_libreoffice(path, timeout)
    if not ok:
        status = "libreoffice_missing" if "not found" in message else "recalc_failed"
        print(json.dumps({
            "status": status,
            "workbook": str(path),
            "total_formulas": total_formulas,
            "total_errors": None,
            "error_summary": {},
            "message": message,
            "recalc_command": cmd,
        }, indent=2))
        return 1

    scanned = scan_errors(path)
    status = "success" if scanned["total_errors"] == 0 else "errors_found"
    print(json.dumps({
        "status": status,
        "workbook": str(path),
        "total_formulas": total_formulas,
        "total_errors": scanned["total_errors"],
        "error_summary": scanned["error_summary"],
        "message": message,
        "recalc_command": cmd,
    }, indent=2))
    return 0 if status == "success" else 1


if __name__ == "__main__":
    raise SystemExit(main())
