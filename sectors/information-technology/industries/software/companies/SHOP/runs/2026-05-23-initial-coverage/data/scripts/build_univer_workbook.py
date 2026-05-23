#!/usr/bin/env python3
import json
import re
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter

RUN_DIR = Path('sectors/software/companies/SHOP/runs/2026-05-23-initial-coverage')
XLSX_PATH = RUN_DIR / 'model.xlsx'
SPEC_PATH = RUN_DIR / 'workboo_new.spec.json'

def clean_color(color_obj):
    if not color_obj:
        return None
    # If color is a Theme or Tint color, openpyxl might not resolve it to RGB.
    # But for our workbook, we set them as RGB hex values.
    if color_obj.type == 'rgb' and color_obj.rgb:
        rgb = color_obj.rgb
        if rgb == '00000000' or rgb == '000000':
            return None
        # ARGB to RGB hex
        return f"#{rgb[-6:]}"
    return None

def get_border_side_style(side):
    if not side or not side.style:
        return None
    style_map = {
        'thin': 1,
        'hair': 2,
        'dotted': 3,
        'dashed': 4,
        'dashDot': 5,
        'dashDotDot': 6,
        'double': 7,
        'medium': 8,
        'mediumDashDot': 10,
        'mediumDashDotDot': 11,
        'thick': 8,  # Map thick to medium since Univer doesn't have thick
    }
    s_val = style_map.get(side.style, 1)
    c_val = clean_color(side.color) or '#000000'
    return {'s': s_val, 'cl': {'rgb': c_val}}

def parse_freeze_panes(freeze_str):
    if not freeze_str or freeze_str == 'A1':
        return None
    # e.g., 'B4' -> split before col B (1 col frozen), before row 4 (3 rows frozen)
    match = re.match(r'^([A-Z]+)([1-9][0-9]*)$', freeze_str)
    if not match:
        return None
    col_str, row_str = match.groups()
    
    # Col string to 0-based index
    col_idx = 0
    for char in col_str:
        col_idx = col_idx * 26 + (ord(char) - ord('A') + 1)
    col_idx -= 1
    
    row_idx = int(row_str) - 1
    
    return {
        'xSplit': col_idx,
        'ySplit': row_idx,
        'startRow': row_idx,
        'startColumn': col_idx
    }

def main():
    print(f"Loading {XLSX_PATH}...")
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=False)
    
    style_registry = {}
    style_counter = 0
    
    sheets_spec = []
    
    # We want to keep the exact sheet order
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"Processing sheet: {name} (dimensions: {ws.max_row}x{ws.max_column})")
        
        # Determine actual bounds (crop trailing empty rows/cols to keep size reasonable)
        max_row = 1
        max_col = 1
        for r in range(1, ws.max_row + 1):
            row_has_data = False
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(r, c)
                if cell.value is not None:
                    row_has_data = True
                    max_col = max(max_col, c)
            if row_has_data:
                max_row = max(max_row, r)
        
        print(f"  Cropped dimensions: {max_row}x{max_col}")
        
        # Build rows array
        rows_data = []
        for r in range(1, max_row + 1):
            row_cells = []
            for c in range(1, max_col + 1):
                cell = ws.cell(r, c)
                val = cell.value
                
                # Check for formula
                is_formula = isinstance(val, str) and val.startswith('=')
                
                # Style properties
                font_style = {}
                if cell.font:
                    if cell.font.bold:
                        font_style['bl'] = 1
                    if cell.font.italic:
                        font_style['it'] = 1
                    if cell.font.size:
                        font_style['fs'] = int(cell.font.size)
                    if cell.font.name:
                        font_style['ff'] = cell.font.name
                    f_color = clean_color(cell.font.color)
                    if f_color:
                        font_style['cl'] = {'rgb': f_color}
                        
                fill_style = {}
                if cell.fill and cell.fill.fill_type == 'solid':
                    bg_color = clean_color(cell.fill.fgColor)
                    if bg_color:
                        fill_style['bg'] = {'rgb': bg_color}
                        
                align_style = {}
                if cell.alignment:
                    h_align = cell.alignment.horizontal
                    if h_align in ['left', 'center', 'right']:
                        align_style['ht'] = {'left': 1, 'center': 2, 'right': 3}[h_align]
                    v_align = cell.alignment.vertical
                    if v_align in ['top', 'center', 'bottom']:
                        align_style['vt'] = {'top': 1, 'center': 2, 'bottom': 3}[v_align]
                        
                border_style = {}
                if cell.border:
                    bd_t = get_border_side_style(cell.border.top)
                    bd_b = get_border_side_style(cell.border.bottom)
                    bd_l = get_border_side_style(cell.border.left)
                    bd_r = get_border_side_style(cell.border.right)
                    bd_obj = {}
                    if bd_t: bd_obj['t'] = bd_t
                    if bd_b: bd_obj['b'] = bd_b
                    if bd_l: bd_obj['l'] = bd_l
                    if bd_r: bd_obj['r'] = bd_r
                    if bd_obj:
                        border_style['bd'] = bd_obj
                        
                # Merge font, fill, alignment, borders to single cell style dict
                cell_style = {}
                cell_style.update(font_style)
                cell_style.update(fill_style)
                cell_style.update(align_style)
                cell_style.update(border_style)
                
                # Handle number format
                num_fmt = cell.number_format
                if num_fmt and num_fmt != 'General':
                    # Clean/standardize common Excel format codes for Univer
                    # E.g. _($* #,##0_);_($* (#,##0);_($* "-"_);_(@_) is converted
                    if '$' in num_fmt or 'mm' in num_fmt or '#' in num_fmt or '0' in num_fmt or '%' in num_fmt:
                        cell_style['n'] = {'pattern': num_fmt}
                
                # Check if cell has any style
                style_id = None
                if cell_style:
                    # Deduplicate styles
                    style_key = json.dumps(cell_style, sort_keys=True)
                    if style_key in style_registry:
                        style_id = style_registry[style_key]
                    else:
                        style_id = f"s{style_counter}"
                        style_registry[style_key] = style_id
                        style_counter += 1
                
                # Construct spec cell object
                if is_formula or style_id is not None:
                    c_obj = {}
                    if is_formula:
                        c_obj['formula'] = val
                    else:
                        # Value only if not formula
                        c_obj['value'] = val
                    
                    if style_id is not None:
                        c_obj['style'] = style_id
                        
                    row_cells.append(c_obj)
                else:
                    # Plain scalar value
                    row_cells.append(val if val is not None else '')
            
            # Trim trailing empty cells in rows to make JSON smaller
            while row_cells and row_cells[-1] == '':
                row_cells.pop()
                
            rows_data.append(row_cells)
            
        # Get columnData (widths)
        column_data = {}
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            dim = ws.column_dimensions.get(col_letter)
            if dim and dim.width:
                # Convert Excel character width to pixels (approx: chars * 7.5 + 10)
                width_px = int(dim.width * 7.5 + 10)
                column_data[str(col_idx - 1)] = {'w': width_px}
            else:
                # Default width
                column_data[str(col_idx - 1)] = {'w': 220 if col_idx == 1 else 130}
                
        # Get rowData (heights)
        row_data = {}
        for r_idx in range(1, max_row + 1):
            dim = ws.row_dimensions.get(r_idx)
            if dim and dim.height:
                row_data[str(r_idx - 1)] = {'h': int(dim.height)}
                
        # Get merges
        merge_data = []
        for rng in ws.merged_cells.ranges:
            # Check if this range falls inside our cropped bounds
            if rng.min_row <= max_row and rng.min_col <= max_col:
                merge_data.append({
                    'startRow': rng.min_row - 1,
                    'endRow': min(rng.max_row, max_row) - 1,
                    'startColumn': rng.min_col - 1,
                    'endColumn': min(rng.max_col, max_col) - 1
                })
                
        # Get freeze pane config
        freeze_config = parse_freeze_panes(ws.freeze_panes)
        
        sheet_spec = {
            'name': name,
            'rows': rows_data,
            'columnCount': max(max_col, 26),
            'rowCount': max(max_row, 100)
        }
        if column_data:
            sheet_spec['columnData'] = column_data
        if row_data:
            sheet_spec['rowData'] = row_data
        if merge_data:
            sheet_spec['mergeData'] = merge_data
        if freeze_config:
            sheet_spec['freeze'] = freeze_config
            
        sheets_spec.append(sheet_spec)
        
    # Reconstruct style map for Univer
    univer_styles = {}
    for style_key, style_id in style_registry.items():
        univer_styles[style_id] = json.loads(style_key)
        
    # Reconstruct outputs
    outputs = [
        {'key':'valuation.implied_share_price','sheet':'DCF','cell':'B107','value':96.30,'unit':'USD/share'},
        {'key':'valuation.upside_downside','sheet':'DCF','cell':'B109','value':-0.0650,'unit':'percent'},
        {'key':'valuation.enterprise_value','sheet':'DCF','cell':'B102','value':119625,'unit':'USD millions'},
        {'key':'valuation.equity_value','sheet':'DCF','cell':'B105','value':125186,'unit':'USD millions'},
        {'key':'market.current_price','sheet':'DCF','cell':'B108','value':103.00,'unit':'USD/share'},
        {'key':'assumption.wacc','sheet':'DCF','cell':'B97','value':0.095,'unit':'percent'},
        {'key':'assumption.terminal_growth','sheet':'DCF','cell':'B98','value':0.035,'unit':'percent'},
        {'key':'financial.fy2026_revenue','sheet':'DCF','cell':'F71','value':14800,'unit':'USD millions'},
        {'key':'financial.fy2035_fcf','sheet':'DCF','cell':'O82','value':10978,'unit':'USD millions'}
    ]
    
    spec_data = {
        'id': 'shop-dcf-2026-05-23',
        'name': 'SHOP DCF Model',
        'styles': univer_styles,
        'sheets': sheets_spec,
        'outputs': outputs
    }
    
    print(f"Writing spec to {SPEC_PATH}...")
    SPEC_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(SPEC_PATH, 'w') as f:
        json.dump(spec_data, f, indent=2)
        f.write('\n')
        
    print("Done!")

if __name__ == '__main__':
    main()
