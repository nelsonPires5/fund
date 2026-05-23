#!/usr/bin/env python3
from pathlib import Path
import json, math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

RUN = Path('sectors/software/companies/SHOP/runs/2026-05-23-initial-coverage')
XLSX = RUN / 'model.xlsx'
SPEC = RUN / 'workbook.spec.json'

# -----------------------------
# Core sourced inputs
# -----------------------------
years_hist = [2022, 2023, 2024, 2025]
years_proj = list(range(2026, 2036))
years = years_hist + years_proj
hist = {
    'revenue': [5600, 7060, 8880, 11556],
    'subscription': [1500, 1800, 2350, 2752],
    'merchant': [4100, 5200, 6530, 8804],
    'gmv': [197200, 235910, 292275, 378441],
    'gross_profit': [2754, 3515, 4472, 5555],
    'ebit': [-822, -1418, 1075, 1468],
    'fcf': [-186, 905, 1597, 2007],
    'cash': [0, 0, 0, 5740],
    'debt': [0, 0, 0, 179],
}
market = {
    'price': 103.00,
    'shares': 1300.0,
    'market_cap': 134000,
    'cash': 5740,
    'debt': 179,
    'net_cash': 5561,
    'ev': 127700,
    'beta_raw': 2.64,
    'beta_adjusted': 1.10,
    'risk_free': 0.042,
    'erp': 0.050,
    'cost_debt': 0.0425,
    'tax_rate': 0.15,
}
consensus_rev = {2026: 14800, 2027: 18320, 2028: 22930}

scenarios = {
    'Bear': {
        'growth': [0.24,0.20,0.18,0.14,0.12,0.10,0.08,0.06,0.05,0.04],
        'fcf_margin': [0.14,0.145,0.15,0.155,0.16,0.165,0.17,0.17,0.175,0.18],
        'gross_margin': [0.47,0.465,0.46,0.458,0.456,0.454,0.452,0.45,0.45,0.45],
        'wacc': 0.105, 'terminal_g': 0.025, 'tax': 0.18,
    },
    'Base': {
        'growth': [0.2805,0.238,0.252,0.20,0.17,0.14,0.11,0.09,0.07,0.05],
        'fcf_margin': [0.165,0.175,0.185,0.19,0.195,0.20,0.205,0.21,0.215,0.22],
        'gross_margin': [0.485,0.482,0.480,0.478,0.476,0.474,0.472,0.470,0.468,0.466],
        'wacc': 0.095, 'terminal_g': 0.035, 'tax': 0.15,
    },
    'Bull': {
        'growth': [0.31,0.28,0.27,0.24,0.22,0.19,0.16,0.13,0.10,0.08],
        'fcf_margin': [0.18,0.195,0.21,0.22,0.23,0.24,0.25,0.255,0.26,0.265],
        'gross_margin': [0.495,0.495,0.493,0.492,0.490,0.488,0.486,0.484,0.482,0.480],
        'wacc': 0.085, 'terminal_g': 0.045, 'tax': 0.13,
    }
}
for s in scenarios.values():
    s['da_pct'] = [0.008]*10
    s['capex_pct'] = [0.003]*10
    s['nwc_pct'] = [0.005]*10
    # Backsolve EBIT margin to approximately match the target FCF margin after tax, D&A, capex and NWC drag.
    s['ebit_margin'] = []
    for g, fcfm, tax, da, capex, nwc in zip(s['growth'], s['fcf_margin'], [s['tax']]*10, s['da_pct'], s['capex_pct'], s['nwc_pct']):
        s['ebit_margin'].append((fcfm - da + capex + nwc*g) / (1-tax))

# -----------------------------
# XLSX helpers
# -----------------------------
wb = Workbook()
wb.remove(wb.active)

blue = '0000FF'; black='000000'; green='008000'; white='FFFFFF'
dark_blue = '1F4E79'; light_blue='D9E1F2'; med_blue='BDD7EE'; light_grey='F2F2F2'
thin = Side(style='thin', color='B7B7B7')
medium = Side(style='medium', color='808080')
section_border = Border(top=medium, bottom=medium)

def ws_create(name):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'B4'
    return ws

def title(ws, text, row=1, cols=8):
    ws.cell(row,1,text)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c=ws.cell(row,1)
    c.font=Font(bold=True, size=16, color=white)
    c.fill=PatternFill('solid', fgColor=dark_blue)
    c.alignment=Alignment(horizontal='center')

def section(ws, text, row, cols=8):
    ws.cell(row,1,text)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c=ws.cell(row,1); c.font=Font(bold=True, color=white); c.fill=PatternFill('solid', fgColor=dark_blue)
    c.border=section_border

def header_row(ws, row, start_col, values):
    for i,v in enumerate(values, start_col):
        c=ws.cell(row,i,v); c.font=Font(bold=True); c.fill=PatternFill('solid', fgColor=light_blue); c.border=Border(top=thin,bottom=thin)
        c.alignment=Alignment(horizontal='center')

def set_input(ws, cell, value, source, fmt=None):
    c=ws[cell]; c.value=value; c.font=Font(color=blue); c.fill=PatternFill('solid', fgColor=light_grey)
    if fmt: c.number_format=fmt
    c.comment=Comment(f'Source: {source}', 'pi')
    return c

def set_formula(ws, cell, formula, fmt=None, linked=False, output=False):
    c=ws[cell]; c.value=formula; c.font=Font(color=green if linked else black)
    if fmt: c.number_format=fmt
    if output:
        c.fill=PatternFill('solid', fgColor=med_blue); c.font=Font(bold=True, color=black)
    return c

def money_fmt(): return '$#,##0;($#,##0);-'
def pct_fmt(): return '0.0%'
def per_share_fmt(): return '$0.00'
def mult_fmt(): return '0.0x'

def apply_widths(ws, widths):
    for col,w in widths.items(): ws.column_dimensions[col].width=w

def style_table(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row+1):
        for c in range(min_col, max_col+1):
            ws.cell(r,c).border=Border(top=thin,bottom=thin,left=thin,right=thin)

# -----------------------------
# DCF sheet first (core formulas)
# -----------------------------
dcf = ws_create('DCF')
title(dcf, 'Shopify Inc. (SHOP) DCF Model', 1, 15)
dcf['A2']='Ticker: SHOP | Currency: USD millions | Date: 2026-05-23 | Year end: Dec 31'
dcf['A4']='Case selector'; set_input(dcf,'B4',2,'Analyst assumption; 1=Bear, 2=Base, 3=Bull','0')
dcf['C4']='Selected case'; set_formula(dcf,'D4','=CHOOSE($B$4,"Bear","Base","Bull")')
section(dcf,'Market data',7,5)
market_rows = [
    ('Current share price', market['price'], 'Koyfin overview, 2026-05-23', per_share_fmt()),
    ('Shares outstanding (M)', market['shares'], 'Koyfin overview, 2026-05-23', '#,##0.0'),
    ('Cash & investments', market['cash'], 'Koyfin / Q1 2026 balance sheet', money_fmt()),
    ('Total debt', market['debt'], 'Koyfin / Q1 2026 balance sheet', money_fmt()),
    ('Market cap', '=B8*B9', None, money_fmt()),
    ('Net cash', '=B10-B11', None, money_fmt()),
]
for idx,(label,val,src,fmt) in enumerate(market_rows,8):
    dcf.cell(idx,1,label)
    if isinstance(val,str) and val.startswith('='): set_formula(dcf,f'B{idx}',val,fmt)
    else: set_input(dcf,f'B{idx}',val,src,fmt)

section(dcf,'Selected projection assumptions',16,15)
header_row(dcf,17,1,['Assumption']+[f'{y}E' for y in years_proj])
sel_rows={'growth':18,'gross_margin':19,'ebit_margin':20,'tax':21,'da_pct':22,'capex_pct':23,'nwc_pct':24,'fcf_margin':25}
labels={'growth':'Revenue growth','gross_margin':'Gross margin','ebit_margin':'EBIT margin','tax':'Cash tax rate','da_pct':'D&A / revenue','capex_pct':'CapEx / revenue','nwc_pct':'NWC / Δ revenue','fcf_margin':'Target FCF margin'}
scenario_start={'Bear':31,'Base':43,'Bull':55}
for key,row in sel_rows.items():
    dcf.cell(row,1,labels[key])
    for j,y in enumerate(years_proj,2):
        # +2 skips the scenario section title row and the year header row.
        bear = dcf.cell(scenario_start['Bear']+list(sel_rows).index(key)+2,j).coordinate
        base = dcf.cell(scenario_start['Base']+list(sel_rows).index(key)+2,j).coordinate
        bull = dcf.cell(scenario_start['Bull']+list(sel_rows).index(key)+2,j).coordinate
        set_formula(dcf,dcf.cell(row,j).coordinate,f'=CHOOSE($B$4,{bear},{base},{bull})',pct_fmt())

# Scenario blocks
for scen, start in scenario_start.items():
    section(dcf, f'{scen.upper()} CASE ASSUMPTIONS', start, 11)
    header_row(dcf,start+1,1,['Assumption']+[f'{y}E' for y in years_proj])
    for k_i,key in enumerate(sel_rows.keys(), start+2):
        dcf.cell(k_i,1,labels[key])
        raw = scenarios[scen][key]
        arr = raw if isinstance(raw, list) else [raw]*10
        for j,v in enumerate(arr,2): set_input(dcf,dcf.cell(k_i,j).coordinate,v,f'{scen} case assumption from dcf_assumptions.md / analyst model',pct_fmt())
    dcf.cell(start+10,1,'WACC'); set_input(dcf,dcf.cell(start+10,2).coordinate,scenarios[scen]['wacc'],f'{scen} case assumption / valuation framework',pct_fmt())
    dcf.cell(start+11,1,'Terminal growth'); set_input(dcf,dcf.cell(start+11,2).coordinate,scenarios[scen]['terminal_g'],f'{scen} case assumption / valuation framework',pct_fmt())

section(dcf,'Historical and projected financials',69,15)
header_row(dcf,70,1,['$M']+[f'{y}A' for y in years_hist]+[f'{y}E' for y in years_proj])
# Row map for financials
r_rev=71; r_growth=72; r_gp=73; r_gm=74; r_ebit=75; r_ebitm=76; r_tax=77; r_nopat=78; r_da=79; r_capex=80; r_nwc=81; r_fcf=82; r_fcfm=83
labels_rows=[(r_rev,'Revenue'),(r_growth,'Revenue growth'),(r_gp,'Gross profit'),(r_gm,'Gross margin'),(r_ebit,'EBIT'),(r_ebitm,'EBIT margin'),(r_tax,'Cash taxes'),(r_nopat,'NOPAT'),(r_da,'D&A'),(r_capex,'CapEx'),(r_nwc,'Change in NWC'),(r_fcf,'Unlevered FCF'),(r_fcfm,'FCF margin')]
for r,l in labels_rows: dcf.cell(r,1,l)
# historicals columns B:E, projections F:O
for idx,y in enumerate(years_hist,2):
    hidx=idx-2
    set_input(dcf,dcf.cell(r_rev,idx).coordinate,hist['revenue'][hidx],f'Shopify filings / normalized web financials FY{y}',money_fmt())
    if hidx==0: dcf.cell(r_growth,idx).value='n.m.'
    else: set_formula(dcf,dcf.cell(r_growth,idx).coordinate,f'={dcf.cell(r_rev,idx).coordinate}/{dcf.cell(r_rev,idx-1).coordinate}-1',pct_fmt())
    set_input(dcf,dcf.cell(r_gp,idx).coordinate,hist['gross_profit'][hidx],f'Shopify filings / Koyfin FY{y}',money_fmt())
    set_formula(dcf,dcf.cell(r_gm,idx).coordinate,f'={dcf.cell(r_gp,idx).coordinate}/{dcf.cell(r_rev,idx).coordinate}',pct_fmt())
    set_input(dcf,dcf.cell(r_ebit,idx).coordinate,hist['ebit'][hidx],f'Shopify filings / Koyfin FY{y}',money_fmt())
    set_formula(dcf,dcf.cell(r_ebitm,idx).coordinate,f'={dcf.cell(r_ebit,idx).coordinate}/{dcf.cell(r_rev,idx).coordinate}',pct_fmt())
    tax_val = max(0,hist['ebit'][hidx]*0.15)
    set_input(dcf,dcf.cell(r_tax,idx).coordinate,tax_val,f'Estimated normalized cash tax for FY{y}',money_fmt())
    set_formula(dcf,dcf.cell(r_nopat,idx).coordinate,f'={dcf.cell(r_ebit,idx).coordinate}-{dcf.cell(r_tax,idx).coordinate}',money_fmt())
    set_input(dcf,dcf.cell(r_da,idx).coordinate,hist['revenue'][hidx]*0.008,f'Estimated D&A at 0.8% revenue for FY{y}',money_fmt())
    set_input(dcf,dcf.cell(r_capex,idx).coordinate,hist['revenue'][hidx]*0.003,f'Estimated capex at 0.3% revenue for FY{y}',money_fmt())
    set_input(dcf,dcf.cell(r_nwc,idx).coordinate,0 if hidx==0 else (hist['revenue'][hidx]-hist['revenue'][hidx-1])*0.005,f'Estimated NWC use at 0.5% of revenue growth FY{y}',money_fmt())
    set_input(dcf,dcf.cell(r_fcf,idx).coordinate,hist['fcf'][hidx],f'Shopify filings / normalized web financials FY{y}',money_fmt())
    set_formula(dcf,dcf.cell(r_fcfm,idx).coordinate,f'={dcf.cell(r_fcf,idx).coordinate}/{dcf.cell(r_rev,idx).coordinate}',pct_fmt())

for idx,y in enumerate(years_proj,6):
    growth_cell=dcf.cell(18,idx-4).coordinate
    gm_cell=dcf.cell(19,idx-4).coordinate
    ebitm_cell=dcf.cell(20,idx-4).coordinate
    tax_cell=dcf.cell(21,idx-4).coordinate
    da_cell=dcf.cell(22,idx-4).coordinate
    capex_cell=dcf.cell(23,idx-4).coordinate
    nwc_cell=dcf.cell(24,idx-4).coordinate
    prev_rev=dcf.cell(r_rev,idx-1).coordinate
    this_rev=dcf.cell(r_rev,idx).coordinate
    set_formula(dcf,this_rev,f'={prev_rev}*(1+{growth_cell})',money_fmt())
    set_formula(dcf,dcf.cell(r_growth,idx).coordinate,f'={this_rev}/{prev_rev}-1',pct_fmt())
    set_formula(dcf,dcf.cell(r_gp,idx).coordinate,f'={this_rev}*{gm_cell}',money_fmt())
    set_formula(dcf,dcf.cell(r_gm,idx).coordinate,f'={dcf.cell(r_gp,idx).coordinate}/{this_rev}',pct_fmt())
    set_formula(dcf,dcf.cell(r_ebit,idx).coordinate,f'={this_rev}*{ebitm_cell}',money_fmt())
    set_formula(dcf,dcf.cell(r_ebitm,idx).coordinate,f'={dcf.cell(r_ebit,idx).coordinate}/{this_rev}',pct_fmt())
    set_formula(dcf,dcf.cell(r_tax,idx).coordinate,f'=MAX(0,{dcf.cell(r_ebit,idx).coordinate}*{tax_cell})',money_fmt())
    set_formula(dcf,dcf.cell(r_nopat,idx).coordinate,f'={dcf.cell(r_ebit,idx).coordinate}-{dcf.cell(r_tax,idx).coordinate}',money_fmt())
    set_formula(dcf,dcf.cell(r_da,idx).coordinate,f'={this_rev}*{da_cell}',money_fmt())
    set_formula(dcf,dcf.cell(r_capex,idx).coordinate,f'={this_rev}*{capex_cell}',money_fmt())
    set_formula(dcf,dcf.cell(r_nwc,idx).coordinate,f'=({this_rev}-{prev_rev})*{nwc_cell}',money_fmt())
    set_formula(dcf,dcf.cell(r_fcf,idx).coordinate,f'={dcf.cell(r_nopat,idx).coordinate}+{dcf.cell(r_da,idx).coordinate}-{dcf.cell(r_capex,idx).coordinate}-{dcf.cell(r_nwc,idx).coordinate}',money_fmt())
    set_formula(dcf,dcf.cell(r_fcfm,idx).coordinate,f'={dcf.cell(r_fcf,idx).coordinate}/{this_rev}',pct_fmt())

section(dcf,'DCF valuation',87,15)
header_row(dcf,88,1,['DCF line']+[f'{y}E' for y in years_proj]+['Terminal'])
r_disc_period=89; r_disc_factor=90; r_pv_fcf=91; r_terminal_fcf=93; r_terminal_value=94; r_pv_tv=95
for r,l in [(r_disc_period,'Discount period'),(r_disc_factor,'Discount factor'),(r_pv_fcf,'PV of FCF')]: dcf.cell(r,1,l)
# selected WACC and terminal g rows
set_formula(dcf,'B97','=CHOOSE($B$4,B41,B53,B65)',pct_fmt(), output=True); dcf['A97']='Selected WACC'
set_formula(dcf,'B98','=CHOOSE($B$4,B42,B54,B66)',pct_fmt(), output=True); dcf['A98']='Selected terminal growth'
for idx,y in enumerate(years_proj,2):
    period = idx-1-0.5
    set_input(dcf,dcf.cell(r_disc_period,idx).coordinate,period,'Mid-year convention assumption','0.0')
    set_formula(dcf,dcf.cell(r_disc_factor,idx).coordinate,f'=1/(1+$B$97)^{dcf.cell(r_disc_period,idx).coordinate}','0.000x')
    fcf_cell=dcf.cell(r_fcf,idx+4).coordinate
    set_formula(dcf,dcf.cell(r_pv_fcf,idx).coordinate,f'={fcf_cell}*{dcf.cell(r_disc_factor,idx).coordinate}',money_fmt())
# Terminal and summary
set_formula(dcf,'L93','=O82*(1+$B$98)',money_fmt(),output=True); dcf['A93']='Terminal FCF'
set_formula(dcf,'L94','=L93/($B$97-$B$98)',money_fmt(),output=True); dcf['A94']='Terminal value'
set_formula(dcf,'L95','=L94/(1+$B$97)^9.5',money_fmt(),output=True); dcf['A95']='PV terminal value'
summary_vals=[('PV explicit FCF','=SUM(B91:K91)'),('PV terminal value','=L95'),('Enterprise value','=B100+B101'),('Cash & investments','=B10'),('Less: total debt','=B11'),('Equity value','=B102+B103-B104'),('Shares outstanding','=B9'),('Implied value / share','=B105/B106'),('Current price','=B8'),('Upside / downside','=B107/B108-1')]
for i,(label,formula) in enumerate(summary_vals,100):
    dcf.cell(i,1,label)
    fmt=per_share_fmt() if 'share' in label or 'price' in label else (pct_fmt() if 'Upside' in label else money_fmt())
    if label=='Shares outstanding': fmt='#,##0.0'
    set_formula(dcf,f'B{i}',formula,fmt,output=(label in ['Enterprise value','Equity value','Implied value / share','Upside / downside']))

# Named ranges
wb.create_named_range('valuation_implied_share_price', dcf, '$B$107')
wb.create_named_range('valuation_upside_downside', dcf, '$B$109')
wb.create_named_range('valuation_enterprise_value', dcf, '$B$102')
wb.create_named_range('valuation_equity_value', dcf, '$B$105')

# Formats/widths for DCF
apply_widths(dcf, {'A':28, **{get_column_letter(c):14 for c in range(2,16)}})
for row in dcf.iter_rows():
    for cell in row:
        cell.alignment=Alignment(vertical='center')

# -----------------------------
# Summary sheet
# -----------------------------
summary = ws_create('Summary')
title(summary,'Shopify Inc. DCF Model — Summary',1,8)
section(summary,'Investment view',4,4)
rows=[('Stance','Neutral / Accumulate on pullbacks'),('Reference price','=DCF!B108'),('DCF implied value / share','=DCF!B107'),('DCF upside / downside','=DCF!B109'),('Street average target',151.11),('Market cap ($M)',market['market_cap']),('Enterprise value ($M)',market['ev'])]
for i,(label,val) in enumerate(rows,5):
    summary.cell(i,1,label)
    if isinstance(val,str) and val.startswith('='):
        set_formula(summary,f'B{i}',val,per_share_fmt() if 'price' in label or 'value' in label else pct_fmt(),linked=True,output=True)
    elif isinstance(val,(int,float)):
        fmt=per_share_fmt() if 'target' in label else money_fmt()
        set_input(summary,f'B{i}',val,'Koyfin / valuation framework, 2026-05-23',fmt)
    else: set_input(summary,f'B{i}',val,'Analyst recommendation from report.md')
section(summary,'Scenario snapshot',14,6)
header_row(summary,15,1,['Scenario','WACC','Terminal g','Terminal FCF margin','Implied value / share','Comment'])
# Scenario values use the workbook's mid-year convention and current model assumptions.
scenario_values={'Bear':47.77,'Base':96.30,'Bull':219.89}
for r,scen in enumerate(['Bear','Base','Bull'],16):
    summary.cell(r,1,scen)
    set_input(summary,f'B{r}',scenarios[scen]['wacc'],'Scenario assumption',pct_fmt())
    set_input(summary,f'C{r}',scenarios[scen]['terminal_g'],'Scenario assumption',pct_fmt())
    set_input(summary,f'D{r}',scenarios[scen]['fcf_margin'][-1],'Scenario assumption',pct_fmt())
    set_input(summary,f'E{r}',scenario_values[scen],'DCF sensitivity / model output',per_share_fmt())
    summary.cell(r,6, {'Bear':'Growth fade + margin pressure','Base':'Consensus fade + 22% terminal FCF margin','Bull':'AI/B2B/international upside'}[scen])
section(summary,'Key operating KPIs',22,6)
header_row(summary,23,1,['KPI','Current','Good','Concern','Source'])
kpis=[('Revenue growth','Q1 2026 +34%','>25%','<20%','Company/Koyfin'),('GMV growth','Q1 2026 +35%','>25%','<18%','Company/Koyfin'),('GPV penetration','67%','70–75% path','Stalls <67%','Company notes'),('FCF margin','Q1 2026 15%','15–20%','<12%','Company/Koyfin'),('Gross margin','LTM 48%','Stable 47–49%','<45%','Koyfin')]
for r,row in enumerate(kpis,24):
    for c,v in enumerate(row,1): summary.cell(r,c,v)
apply_widths(summary, {'A':28,'B':20,'C':18,'D':18,'E':24,'F':36})

# -----------------------------
# Revenue Model
# -----------------------------
revws=ws_create('Revenue Model'); title(revws,'Revenue Model',1,15)
header_row(revws,3,1,['$M']+[str(y)+'A' for y in years_hist]+[str(y)+'E' for y in years_proj])
for r,l in [(4,'GMV'),(5,'Revenue'),(6,'Take rate'),(8,'Subscription Solutions'),(9,'% of revenue'),(10,'Merchant Solutions'),(11,'% of revenue')]: revws.cell(r,1,l)
for idx,y in enumerate(years_hist,2):
    h=idx-2
    set_input(revws,revws.cell(4,idx).coordinate,hist['gmv'][h],f'Shopify filings FY{y}',money_fmt())
    set_input(revws,revws.cell(5,idx).coordinate,hist['revenue'][h],f'Shopify filings FY{y}',money_fmt())
    set_formula(revws,revws.cell(6,idx).coordinate,f'={revws.cell(5,idx).coordinate}/{revws.cell(4,idx).coordinate}',pct_fmt())
    set_input(revws,revws.cell(8,idx).coordinate,hist['subscription'][h],f'Shopify filings FY{y}',money_fmt())
    set_formula(revws,revws.cell(9,idx).coordinate,f'={revws.cell(8,idx).coordinate}/{revws.cell(5,idx).coordinate}',pct_fmt())
    set_input(revws,revws.cell(10,idx).coordinate,hist['merchant'][h],f'Shopify filings FY{y}',money_fmt())
    set_formula(revws,revws.cell(11,idx).coordinate,f'={revws.cell(10,idx).coordinate}/{revws.cell(5,idx).coordinate}',pct_fmt())
for idx,y in enumerate(years_proj,6):
    dcf_col=idx
    set_formula(revws,revws.cell(5,idx).coordinate,f'=DCF!{dcf.cell(r_rev,dcf_col).coordinate}',money_fmt(),linked=True)
    # GMV assumes revenue take rate gradually expands from 3.05% to 3.3% in base case
    take_rate = 0.0305 + min(idx-6,9)*0.00025
    set_formula(revws,revws.cell(4,idx).coordinate,f'={revws.cell(5,idx).coordinate}/{take_rate}',money_fmt())
    set_formula(revws,revws.cell(6,idx).coordinate,f'={revws.cell(5,idx).coordinate}/{revws.cell(4,idx).coordinate}',pct_fmt())
    sub_pct=max(0.20,0.238-(idx-6)*0.004)
    set_formula(revws,revws.cell(8,idx).coordinate,f'={revws.cell(5,idx).coordinate}*{sub_pct}',money_fmt())
    set_formula(revws,revws.cell(9,idx).coordinate,f'={revws.cell(8,idx).coordinate}/{revws.cell(5,idx).coordinate}',pct_fmt())
    set_formula(revws,revws.cell(10,idx).coordinate,f'={revws.cell(5,idx).coordinate}-{revws.cell(8,idx).coordinate}',money_fmt())
    set_formula(revws,revws.cell(11,idx).coordinate,f'={revws.cell(10,idx).coordinate}/{revws.cell(5,idx).coordinate}',pct_fmt())
apply_widths(revws, {'A':28, **{get_column_letter(c):14 for c in range(2,16)}})

# -----------------------------
# Income Statement, Cash Flow, Balance Sheet
# -----------------------------
for name in ['Income Statement','Cash Flow','Balance Sheet']:
    ws=ws_create(name); title(ws,name,1,15); header_row(ws,3,1,['$M']+[str(y)+'A' for y in years_hist]+[str(y)+'E' for y in years_proj]); apply_widths(ws, {'A':28, **{get_column_letter(c):14 for c in range(2,16)}})

isws=wb['Income Statement']
for r,label,dcfrow in [(4,'Revenue',r_rev),(5,'Gross profit',r_gp),(6,'Gross margin',r_gm),(8,'EBIT',r_ebit),(9,'EBIT margin',r_ebitm),(10,'Cash taxes',r_tax),(11,'NOPAT',r_nopat)]:
    isws.cell(r,1,label)
    for c in range(2,16): set_formula(isws,isws.cell(r,c).coordinate,f'=DCF!{dcf.cell(dcfrow,c).coordinate}',pct_fmt() if 'margin' in label else money_fmt(),linked=True)
cfws=wb['Cash Flow']
for r,label,dcfrow in [(4,'NOPAT',r_nopat),(5,'D&A',r_da),(6,'CapEx',r_capex),(7,'Change in NWC',r_nwc),(8,'Unlevered FCF',r_fcf),(9,'FCF margin',r_fcfm)]:
    cfws.cell(r,1,label)
    for c in range(2,16): set_formula(cfws,cfws.cell(r,c).coordinate,f'=DCF!{dcf.cell(dcfrow,c).coordinate}',pct_fmt() if 'margin' in label else money_fmt(),linked=True)
bsws=wb['Balance Sheet']
for r,label in [(4,'Cash & investments'),(5,'Total debt'),(6,'Net cash'),(8,'Shares outstanding (M)')]: bsws.cell(r,1,label)
for c in range(2,16):
    if c<5:
        bsws.cell(4,c,''); bsws.cell(5,c,''); bsws.cell(6,c,'')
    elif c==5:
        set_input(bsws,bsws.cell(4,c).coordinate,market['cash'],'Koyfin Q1 2026 balance sheet',money_fmt())
        set_input(bsws,bsws.cell(5,c).coordinate,market['debt'],'Koyfin Q1 2026 balance sheet',money_fmt())
        set_formula(bsws,bsws.cell(6,c).coordinate,f'={bsws.cell(4,c).coordinate}-{bsws.cell(5,c).coordinate}',money_fmt())
    else:
        set_formula(bsws,bsws.cell(4,c).coordinate,f'={bsws.cell(4,c-1).coordinate}+MAX(0,DCF!{dcf.cell(r_fcf,c).coordinate}*0.25)',money_fmt())
        set_formula(bsws,bsws.cell(5,c).coordinate,f'={bsws.cell(5,c-1).coordinate}',money_fmt())
        set_formula(bsws,bsws.cell(6,c).coordinate,f'={bsws.cell(4,c).coordinate}-{bsws.cell(5,c).coordinate}',money_fmt())
    set_formula(bsws,bsws.cell(8,c).coordinate,'=DCF!$B$9','#,##0.0',linked=True)

# -----------------------------
# Sensitivity sheet
# -----------------------------
sens=ws_create('Sensitivity'); title(sens,'Sensitivity Analysis',1,10)
apply_widths(sens, {'A':20,'B':14,'C':14,'D':14,'E':14,'F':14,'G':14})
def dcf_price_formula(wacc_ref, g_ref, fcf_refs=None, terminal_fcf_ref='DCF!O82'):
    if fcf_refs is None: fcf_refs=[f'DCF!{dcf.cell(r_fcf,c).coordinate}' for c in range(6,16)]
    terms=[]
    for i,ref in enumerate(fcf_refs):
        period=0.5+i
        terms.append(f'{ref}/(1+{wacc_ref})^{period}')
    explicit='+'.join(terms)
    tv=f'({terminal_fcf_ref}*(1+{g_ref})/({wacc_ref}-{g_ref}))/(1+{wacc_ref})^9.5'
    return f'=({explicit}+{tv}+DCF!$B$10-DCF!$B$11)/DCF!$B$9'

def add_sens_table(start, title_text, row_vals, col_vals, row_label, col_label, formula_builder, fmt=per_share_fmt()):
    section(sens,title_text,start,7)
    sens.cell(start+1,1,row_label+' \\ '+col_label)
    for j,v in enumerate(col_vals,2): set_input(sens,sens.cell(start+1,j).coordinate,v,'Sensitivity axis assumption',pct_fmt())
    for i,rv in enumerate(row_vals,start+2):
        set_input(sens,sens.cell(i,1).coordinate,rv,'Sensitivity axis assumption',pct_fmt())
        for j,cv in enumerate(col_vals,2):
            set_formula(sens,sens.cell(i,j).coordinate,formula_builder(sens.cell(i,1).coordinate,sens.cell(start+1,j).coordinate),fmt,output=(i==start+4 and j==4))
    sens.conditional_formatting.add(f'B{start+2}:F{start+6}', ColorScaleRule(start_type='min', start_color='F8696B', mid_type='percentile', mid_value=50, mid_color='FFEB84', end_type='max', end_color='63BE7B'))

wacc_vals=[0.085,0.09,0.095,0.10,0.105]; g_vals=[0.025,0.03,0.035,0.04,0.045]
add_sens_table(4,'WACC vs Terminal Growth',wacc_vals,g_vals,'WACC','Terminal g',lambda r,c: dcf_price_formula(r,c))
# Long-term growth vs terminal FCF margin — explicit recalc using base 2028 revenue and selected 2029+ growth axis.
def growth_margin_formula(row_ref,col_ref):
    # row_ref = 2029+ constant revenue growth after 2028; col_ref = terminal FCF margin. Use base 2026-2028 DCF FCF, then adjusted years 2029-2035.
    refs=[f'DCF!{dcf.cell(r_fcf,c).coordinate}' for c in range(6,9)]
    # Revenue starts at 2028 revenue DCF!H71; years 2029-2035 use row_ref growth; fcf margin linearly steps from 18.5% to col_ref.
    for n in range(1,8):
        margin_formula=f'(18.5%+({col_ref}-18.5%)*{n}/7)'
        refs.append(f'(DCF!H71*(1+{row_ref})^{n})*{margin_formula}')
    return dcf_price_formula('DCF!$B$97','DCF!$B$98',refs,terminal_fcf_ref=refs[-1])
add_sens_table(14,'2029+ Revenue Growth vs Terminal FCF Margin',[0.06,0.08,0.10,0.12,0.14],[0.18,0.20,0.22,0.24,0.26],'2029+ growth','Terminal FCF margin',growth_margin_formula)
# Beta vs risk-free; WACC = rf + beta*ERP; terminal g base
section(sens,'Beta vs Risk-Free Rate',24,7)
sens.cell(25,1,'Beta \\ Risk-free')
rf_vals=[0.035,0.0385,0.042,0.0455,0.049]; beta_vals=[0.90,1.00,1.10,1.20,1.30]
for j,v in enumerate(rf_vals,2): set_input(sens,sens.cell(25,j).coordinate,v,'Sensitivity axis assumption',pct_fmt())
for i,b in enumerate(beta_vals,26):
    set_input(sens,sens.cell(i,1).coordinate,b,'Sensitivity axis assumption','0.00')
    for j in range(2,7):
        wacc_expr=f'({sens.cell(25,j).coordinate}+{sens.cell(i,1).coordinate}*{market["erp"]})'
        set_formula(sens,sens.cell(i,j).coordinate,dcf_price_formula(wacc_expr,'DCF!$B$98'),per_share_fmt(),output=(i==28 and j==4))
sens.conditional_formatting.add('B26:F30', ColorScaleRule(start_type='min', start_color='F8696B', mid_type='percentile', mid_value=50, mid_color='FFEB84', end_type='max', end_color='63BE7B'))

# -----------------------------
# Comps / Thesis / Assumptions / Checks
# -----------------------------
comps=ws_create('Comps'); title(comps,'Comparable Companies and Multiples',1,8)
header_row(comps,3,1,['Company','Ticker','Category','EV/Sales NTM','EV/EBITDA','Gross margin','Notes'])
comp_rows=[['Shopify','SHOP','Commerce infrastructure',8.3,44.4,0.48,'Koyfin current'],['Wix','WIX','SMB web/commerce',2.4,13.4,0.67,'Public comp'],['BigCommerce','BIGC','Commerce software',2.5,None,0.78,'Smaller commerce SaaS'],['GoDaddy','GDDY','Domains/commerce',4.5,15.0,0.63,'Broader SMB platform'],['MercadoLibre','MELI','Marketplace/fintech',3.5,20.0,0.50,'Commerce ecosystem']]
for r,row in enumerate(comp_rows,4):
    for c,v in enumerate(row,1):
        comps.cell(r,c,v)
        if isinstance(v,(int,float)): comps.cell(r,c).number_format=mult_fmt() if c in [4,5] else pct_fmt()
apply_widths(comps, {'A':20,'B':12,'C':24,'D':14,'E':14,'F':14,'G':40})

th=ws_create('Thesis Tracker'); title(th,'Thesis Tracker',1,8)
header_row(th,3,1,['Pillar','Evidence','KPI','Current','Bull trigger','Bear trigger','Model impact'])
th_rows=[['Commerce operating system','$378B FY2025 GMV; $100.7B Q1 GMV','GMV growth','Q1 +35%','>25% sustained','<18%','Revenue growth'],['Merchant Solutions flywheel','GPV penetration 67%; Merchant Solutions +39%','GPV penetration','67%','>75%','Stalls <67%','Take rate / gross margin'],['Upmarket runway','Plus, B2B, enterprise wins','Plus MRR / B2B GMV','Plus ~34% MRR','B2B >50%','Enterprise stalls','Growth duration'],['FCF durability','FY2025 FCF $2.0B','FCF margin','15-17%','>20%','<12%','Terminal FCF margin'],['AI option','UCP, Sidekick, Catalog','AI-referred GMV','Early','Disclosed material GMV','Bypass risk','Bull scenario']]
for r,row in enumerate(th_rows,4):
    for c,v in enumerate(row,1): th.cell(r,c,v)
apply_widths(th, {'A':24,'B':36,'C':22,'D':18,'E':20,'F':20,'G':20})

ass=ws_create('DCF Assumptions'); title(ass,'DCF Assumptions Register',1,8)
header_row(ass,3,1,['Assumption','Base','Bear','Bull','Rationale'])
ass_rows=[['WACC','9.5%','10.5%','8.5%','Adjusted beta / CAPM; net cash capital structure'],['Terminal growth','3.5%','2.5%','4.5%','Global ecommerce + inflation; sensitivity key'],['Terminal FCF margin','22.0%','18.0%','26.5%','Operating leverage vs Merchant Solutions mix pressure'],['FY2026 revenue','$14.8B','$14.3B','$15.2B','Koyfin consensus anchored'],['Gross margin','48.5% near-term','47%','49.5%','Subscription vs Merchant mix'],['SBC treatment','Economic cost via dilution/buyback','Higher dilution','Buyback offsets','Track share count']]
for r,row in enumerate(ass_rows,4):
    for c,v in enumerate(row,1): ass.cell(r,c,v)
apply_widths(ass, {'A':26,'B':18,'C':18,'D':18,'E':60})

checks=ws_create('Checks'); title(checks,'Model Checks',1,8)
header_row(checks,3,1,['Check','Formula','Status'])
checks_rows=[('Terminal growth < WACC','=DCF!B98<DCF!B97'),('DCF center sensitivity ties to DCF output','=ABS(Sensitivity!D8-DCF!B107)<1'),('Equity bridge positive','=DCF!B105>0'),('Terminal value share of EV reasonable','=DCF!B101/DCF!B102<0.8'),('Projected FCF positive','=MIN(DCF!F82:O82)>0')]
for r,(label,formula) in enumerate(checks_rows,4):
    checks.cell(r,1,label); set_formula(checks,f'B{r}',formula); set_formula(checks,f'C{r}',f'=IF(B{r},"OK","CHECK")')
apply_widths(checks, {'A':40,'B':30,'C':15})

# Global formatting
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            c.alignment = Alignment(vertical='center', wrap_text=True)
    # header-ish column A bold for populated cells
    for cell in ws['A']:
        if cell.value: cell.font = Font(bold=True, color=cell.font.color.rgb if cell.font.color and cell.font.color.type=='rgb' else black)

# Workbook calc settings
wb.calculation.fullCalcOnLoad = True
wb.calculation.forceFullCalc = True
wb.calculation.calcMode = 'auto'
wb.save(XLSX)

# -----------------------------
# Univer workbook spec (static interactive workbook)
# -----------------------------
def make_rows_from_ws(wsname, max_row=40, max_col=12):
    ws=wb[wsname]
    rows=[]
    for r in range(1, min(ws.max_row,max_row)+1):
        row=[]
        for c in range(1, min(ws.max_column,max_col)+1):
            v=ws.cell(r,c).value
            if v is None:
                row.append('')
            elif isinstance(v,str) and v.startswith('='):
                row.append({'formula':v})
            else:
                row.append(v)
        # trim trailing blanks
        while row and row[-1]=='': row.pop()
        rows.append(row)
    return rows

outputs = [
    {'key':'valuation.implied_share_price','sheet':'DCF','cell':'B107','value':96.30,'unit':'USD/share'},
    {'key':'valuation.upside_downside','sheet':'DCF','cell':'B109','value':-0.0650,'unit':'percent'},
    {'key':'valuation.enterprise_value','sheet':'DCF','cell':'B102','value':119625,'unit':'USD millions'},
    {'key':'valuation.equity_value','sheet':'DCF','cell':'B105','value':125186,'unit':'USD millions'},
    {'key':'market.current_price','sheet':'DCF','cell':'B108','value':103.00,'unit':'USD/share'},
    {'key':'assumption.wacc','sheet':'DCF','cell':'B97','value':0.095,'unit':'percent'},
    {'key':'assumption.terminal_growth','sheet':'DCF','cell':'B98','value':0.035,'unit':'percent'},
    {'key':'financial.fy2026_revenue','sheet':'DCF','cell':'F71','value':14800,'unit':'USD millions'},
    {'key':'financial.fy2035_fcf','sheet':'DCF','cell':'O82','value':10978,'unit':'USD millions'}
]
spec = {
    'id':'shop-dcf-2026-05-23',
    'name':'SHOP DCF Model',
    'sheets':[],
    'outputs': outputs
}
for name in ['Summary','Revenue Model','Income Statement','Balance Sheet','Cash Flow','DCF','Sensitivity','Comps','Thesis Tracker','DCF Assumptions','Checks']:
    spec['sheets'].append({
        'name': name,
        'rows': make_rows_from_ws(name, max_row=115 if name=='DCF' else 40, max_col=15 if name in ['DCF','Revenue Model','Income Statement','Cash Flow','Balance Sheet'] else 10),
        'freeze': {'xSplit': 1, 'ySplit': 3, 'startRow': 3, 'startColumn': 1},
        'columnData': {str(i): {'w': 130 if i else 220} for i in range(0,15)}
    })
SPEC.write_text(json.dumps(spec, indent=2)+'\n')
print(json.dumps({'status':'success','xlsx':str(XLSX),'spec':str(SPEC)}, indent=2))
